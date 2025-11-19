"""Admin panel routes."""
from flask import render_template, redirect, url_for, flash, request, session, jsonify, send_file, current_app
from flask_login import login_required, current_user, login_user, logout_user
from app.blueprints.admin import admin_bp
from app.models.admin import Administrador
from app.models.user import User
from app.models.product import Producto
from app.models.order import Compra
from app.models.notification import Notificacion
from app.models.visit import VisitaPais, VisitaPersona
from app.models.categoria import Categoria, Subcategoria
from app.models.setting import Slide, Banner, Plantilla, Cabecera
from app.models.comercio import Comercio
from app.models.comment import Comentario
from app.models.coupon import Cupon
from app.models.wishlist import Deseo
from app.models.message import Mensaje
from app.extensions import db, bcrypt
from werkzeug.utils import secure_filename
from functools import wraps
from datetime import datetime
import os
import json
import io
from PIL import Image


def admin_required(f):
    """Decorator to require admin access."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if admin is logged in
        if 'admin_id' not in session:
            flash('Debe iniciar sesión como administrador.', 'error')
            return redirect(url_for('admin.login'))

        # Verify admin still exists and is active
        admin = Administrador.query.get(session['admin_id'])
        if not admin or not admin.is_active_user():
            session.pop('admin_id', None)
            flash('Sesión de administrador inválida.', 'error')
            return redirect(url_for('admin.login'))

        return f(*args, **kwargs)
    return decorated_function


@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login page."""
    # If already logged in as admin, redirect to dashboard
    if 'admin_id' in session:
        return redirect(url_for('admin.dashboard'))

    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        if not email or not password:
            flash('Por favor complete todos los campos.', 'error')
            return render_template('admin/login.html')

        # Find admin user
        admin = Administrador.query.filter_by(email=email).first()

        if admin and admin.check_password(password):
            if admin.is_active_user():
                # Store admin ID in session
                session['admin_id'] = admin.id
                session['admin_email'] = admin.email
                session['admin_nombre'] = admin.nombre
                session['admin_perfil'] = admin.perfil

                flash(f'Bienvenido {admin.nombre}!', 'success')
                return redirect(url_for('admin.dashboard'))
            else:
                flash('Su cuenta de administrador está inactiva.', 'error')
        else:
            flash('Email o contraseña incorrectos.', 'error')

    return render_template('admin/login.html')


@admin_bp.route('/logout')
def logout():
    """Admin logout."""
    session.pop('admin_id', None)
    session.pop('admin_email', None)
    session.pop('admin_nombre', None)
    session.pop('admin_perfil', None)
    flash('Sesión cerrada correctamente.', 'success')
    return redirect(url_for('admin.login'))


@admin_bp.route('/')
@admin_required
def dashboard():
    """Admin dashboard."""
    # Get statistics
    total_users = User.query.count()
    total_products = Producto.query.filter_by(estado=1).count()
    total_orders = Compra.query.count()
    total_visits = VisitaPersona.get_total_visits()

    # Get notifications
    notifications = Notificacion.get_counters()

    # Recent orders
    recent_orders = Compra.query.order_by(Compra.fecha.desc()).limit(10).all()

    # Top products
    top_products = Producto.query.filter_by(estado=1).order_by(Producto.ventas.desc()).limit(5).all()

    # Chart data - Sales by day (last 7 days)
    from datetime import timedelta
    from sqlalchemy import func, cast, Date
    today = datetime.now().date()
    week_ago = today - timedelta(days=6)
    
    sales_by_day = db.session.query(
        cast(Compra.fecha, Date).label('date'),
        func.count(Compra.id).label('count'),
        func.sum(Compra.pago).label('total')
    ).filter(Compra.fecha >= week_ago).group_by(cast(Compra.fecha, Date)).all()
    
    # Chart data - Visits by country (top 5)
    top_countries = VisitaPais.query.order_by(VisitaPais.cantidad.desc()).limit(5).all()
    
    return render_template('admin/dashboard.html',
                         total_users=total_users,
                         total_products=total_products,
                         total_orders=total_orders,
                         total_visits=total_visits,
                         notifications=notifications,
                         recent_orders=recent_orders,
                         top_products=top_products,
                         sales_by_day=sales_by_day,
                         top_countries=top_countries)


@admin_bp.route('/users')
@admin_required
def users():
    """Manage users."""
    # Get statistics for cards
    total_users = User.query.count()
    verified_users = User.query.filter_by(verificacion=0).count()
    pending_users = User.query.filter_by(verificacion=1).count()

    return render_template('admin/users.html',
                         total_users=total_users,
                         verified_users=verified_users,
                         pending_users=pending_users)


@admin_bp.route('/users/ajax')
@admin_required
def users_ajax():
    """AJAX endpoint for DataTables user listing."""
    # DataTables parameters
    draw = request.args.get('draw', type=int, default=1)
    start = request.args.get('start', type=int, default=0)
    length = request.args.get('length', type=int, default=10)
    search_value = request.args.get('search[value]', default='')
    order_column = request.args.get('order[0][column]', type=int, default=0)
    order_dir = request.args.get('order[0][dir]', default='desc')

    # Custom filters
    verificacion_filter = request.args.get('verificacion_filter', default='')

    # Base query
    query = User.query

    # Search filter
    if search_value:
        query = query.filter(
            db.or_(
                User.nombre.like(f'%{search_value}%'),
                User.email.like(f'%{search_value}%'),
                User.telefono.like(f'%{search_value}%') if hasattr(User, 'telefono') else False
            )
        )

    # Verification filter
    if verificacion_filter != '':
        query = query.filter_by(verificacion=int(verificacion_filter))

    # Total records
    total_records = User.query.count()
    filtered_records = query.count()

    # Ordering
    columns = [
        User.id,
        User.nombre,
        User.email,
        User.modo,
        User.fecha,
        User.verificacion
    ]

    if 0 <= order_column < len(columns):
        order_col = columns[order_column]
        if order_dir == 'asc':
            query = query.order_by(order_col.asc())
        else:
            query = query.order_by(order_col.desc())
    else:
        query = query.order_by(User.fecha.desc())

    # Get paginated data
    users = query.offset(start).limit(length).all()

    # Format data for DataTables
    data = []
    for user in users:
        # User column with avatar
        inicial = user.nombre[0].upper() if user.nombre else 'U'
        user_html = f'''<div class="d-flex align-items-center">
            <div class="avatar-circle bg-primary text-white me-2">{inicial}</div>
            <div>
                <strong>{user.nombre}</strong>'''
        if hasattr(user, 'pais') and user.pais:
            user_html += f'<br><small class="text-muted"><i class="fas fa-map-marker-alt"></i> {user.pais}</small>'
        user_html += '</div></div>'

        # Contact column
        contact_html = f'<div><i class="fas fa-envelope text-muted me-1"></i><small>{user.email}</small></div>'
        if hasattr(user, 'telefono') and user.telefono:
            contact_html += f'<div><i class="fas fa-phone text-muted me-1"></i><small>{user.telefono}</small></div>'

        # Mode badge
        if user.modo == 'directo':
            modo_html = '<span class="badge bg-primary"><i class="fas fa-envelope"></i> Directo</span>'
        elif user.modo == 'google':
            modo_html = '<span class="badge bg-danger"><i class="fab fa-google"></i> Google</span>'
        elif user.modo == 'facebook':
            modo_html = '<span class="badge bg-info"><i class="fab fa-facebook"></i> Facebook</span>'
        else:
            modo_html = f'<span class="badge bg-secondary">{user.modo}</span>'

        data.append({
            'id': user.id,
            'usuario': user_html,
            'contacto': contact_html,
            'modo': modo_html,
            'registro': user.fecha.strftime('%d/%m/%Y') if user.fecha else 'N/A',
            'verificacion': user.verificacion,
            'actions': ''  # Will be rendered by DataTables
        })

    return jsonify({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })



@admin_bp.route('/users/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_user(id):
    """Toggle user verification status."""
    try:
        user = User.query.get_or_404(id)
        user.verificacion = 0 if user.verificacion == 1 else 1
        db.session.commit()
        return jsonify({'success': True, 'verificacion': user.verificacion})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/users/<int:id>/orders')
@admin_required
def user_orders(id):
    """View user's order history."""
    user = User.query.get_or_404(id)
    page = request.args.get('page', 1, type=int)
    orders = Compra.query.filter_by(id_usuario=id).order_by(Compra.fecha.desc()).paginate(
        page=page, per_page=25, error_out=False
    )

    return render_template('admin/user_orders.html', user=user, orders=orders)


@admin_bp.route('/users/<int:id>/detail')
@admin_required
def user_detail(id):
    """View user details."""
    from sqlalchemy import func
    user = User.query.get_or_404(id)

    # Get user statistics
    total_orders = Compra.query.filter_by(id_usuario=id).count()
    total_spent = db.session.query(func.sum(Compra.pago)).filter_by(id_usuario=id).scalar() or 0
    total_comments = Comentario.query.filter_by(id_usuario=id).count()
    total_wishlist = Deseo.query.filter_by(id_usuario=id).count()

    # Get recent orders
    recent_orders = Compra.query.filter_by(id_usuario=id).order_by(Compra.fecha.desc()).limit(5).all()

    return render_template('admin/user_detail.html',
                         user=user,
                         total_orders=total_orders,
                         total_spent=total_spent,
                         total_comments=total_comments,
                         total_wishlist=total_wishlist,
                         recent_orders=recent_orders)


@admin_bp.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_user(id):
    """Edit user from admin."""
    from werkzeug.utils import secure_filename
    import os

    user = User.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # Update basic fields
            user.nombre = request.form.get('nombre', '').strip()
            email_nuevo = request.form.get('email', '').strip()

            # Check if email changed and is unique
            if email_nuevo != user.email:
                existing = User.query.filter_by(email=email_nuevo).first()
                if existing:
                    flash('El email ya está registrado por otro usuario.', 'error')
                    return render_template('admin/user_form.html', user=user, edit_mode=True)
                user.email = email_nuevo

            # Handle photo upload
            if 'foto' in request.files:
                file = request.files['foto']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"user_{user.id}_{timestamp}_{filename}"

                    upload_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'app/static/uploads'), 'users')
                    os.makedirs(upload_folder, exist_ok=True)
                    filepath = os.path.join(upload_folder, filename)

                    # Save and resize image
                    from PIL import Image
                    img = Image.open(file)
                    img.thumbnail((400, 400))
                    img.save(filepath, optimize=True, quality=85)

                    user.foto = f"uploads/users/{filename}"

            # Change password if provided
            new_password = request.form.get('new_password', '').strip()
            if new_password:
                user.set_password(new_password)

            db.session.commit()
            flash(f'Usuario "{user.nombre}" actualizado exitosamente.', 'success')
            return redirect(url_for('admin.users'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar usuario: {e}', 'error')

    return render_template('admin/user_form.html', user=user, edit_mode=True)


@admin_bp.route('/users/delete/<int:id>', methods=['POST'])
@admin_required
def delete_user(id):
    """Delete user."""
    try:
        user = User.query.get_or_404(id)

        # Check if user has orders
        orders_count = Compra.query.filter_by(id_usuario=id).count()
        if orders_count > 0:
            flash(f'No se puede eliminar al usuario "{user.nombre}" porque tiene {orders_count} compra(s) asociada(s). Se recomienda desactivar la cuenta en su lugar.', 'error')
            return redirect(url_for('admin.users'))

        # Delete user (wishlist and comments will be deleted by cascade)
        db.session.delete(user)
        db.session.commit()

        flash(f'Usuario "{user.nombre}" eliminado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar usuario: {e}', 'error')

    return redirect(url_for('admin.users'))


@admin_bp.route('/products')
@admin_required
def products():
    """Manage products."""
    categorias = Categoria.query.all()
    return render_template('admin/products.html', categorias=categorias)


@admin_bp.route('/products/ajax')
@admin_required
def products_ajax():
    """AJAX endpoint for DataTables product listing."""
    # DataTables parameters
    draw = request.args.get('draw', type=int, default=1)
    start = request.args.get('start', type=int, default=0)
    length = request.args.get('length', type=int, default=10)
    search_value = request.args.get('search[value]', default='')
    order_column = request.args.get('order[0][column]', type=int, default=0)
    order_dir = request.args.get('order[0][dir]', default='desc')

    # Category filter from custom parameter
    categoria_filter = request.args.get('categoria_filter', default='')

    # Base query
    query = Producto.query

    # Search filter
    if search_value:
        query = query.filter(
            db.or_(
                Producto.titulo.like(f'%{search_value}%'),
                Producto.descripcion.like(f'%{search_value}%'),
                Producto.ruta.like(f'%{search_value}%')
            )
        )

    # Category filter
    if categoria_filter:
        query = query.filter_by(id_categoria=int(categoria_filter))

    # Total records
    total_records = Producto.query.count()
    filtered_records = query.count()

    # Ordering
    columns = [
        Producto.id,
        Producto.portada,
        Producto.titulo,
        Producto.id_categoria,
        Producto.precio,
        Producto.stock,
        Producto.ventas,
        Producto.estado
    ]

    if 0 <= order_column < len(columns):
        order_col = columns[order_column]
        if order_dir == 'asc':
            query = query.order_by(order_col.asc())
        else:
            query = query.order_by(order_col.desc())
    else:
        query = query.order_by(Producto.fecha.desc())

    # Get paginated data
    productos = query.offset(start).limit(length).all()

    # Format data for DataTables
    data = []
    for producto in productos:
        # Get category name
        cat_name = producto.categoria.categoria if producto.categoria else 'N/A'

        # Format price
        precio_html = f'${producto.precio:.2f}'
        if producto.descuento > 0:
            precio_html += f'<br><span class="badge bg-warning text-dark">-{producto.descuento}%</span>'

        # Format stock
        if producto.is_virtual():
            stock_html = '<span class="badge bg-info">Virtual</span>'
        elif producto.stock > 10:
            stock_html = f'<span class="badge bg-success">{producto.stock}</span>'
        elif producto.stock > 0:
            stock_html = f'<span class="badge bg-warning text-dark">{producto.stock}</span>'
        else:
            stock_html = '<span class="badge bg-danger">Agotado</span>'

        # Image HTML
        if producto.portada:
            img_html = f'<img src="/static/{producto.portada}" alt="{producto.titulo}" style="width: 60px; height: 40px; object-fit: cover; border-radius: 4px;">'
        else:
            img_html = '<div style="width: 60px; height: 40px; background: #ddd; display: flex; align-items: center; justify-content: center; border-radius: 4px;"><i class="fas fa-image text-muted"></i></div>'

        data.append({
            'id': producto.id,
            'imagen': img_html,
            'titulo': f'<strong>{producto.titulo}</strong><br><small class="text-muted">{producto.ruta}</small>',
            'categoria': cat_name,
            'precio': precio_html,
            'stock': stock_html,
            'ventas': producto.ventas,
            'estado': producto.estado,
            'actions': ''  # Will be rendered by DataTables
        })

    return jsonify({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@admin_bp.route('/products/create', methods=['GET', 'POST'])
@admin_required
def create_product():
    """Create new product."""
    if request.method == 'POST':
        try:
            # Get form data
            titulo = request.form.get('titulo')
            ruta = request.form.get('ruta')
            id_categoria = request.form.get('id_categoria', type=int)
            id_subcategoria = request.form.get('id_subcategoria', type=int) or None
            tipo = request.form.get('tipo', 'fisico')
            precio = float(request.form.get('precio', 0))
            stock = int(request.form.get('stock', 0)) if tipo == 'fisico' else 0
            descripcion = request.form.get('descripcion', '')
            titular = request.form.get('titular', '')

            # Handle image upload
            portada = ''
            if 'portada' in request.files:
                file = request.files['portada']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    upload_folder = os.path.join('app/static/uploads/productos')
                    if not os.path.exists(upload_folder):
                        os.makedirs(upload_folder)

                    filepath = os.path.join(upload_folder, filename)

                    # Resize image to 1280x720
                    img = Image.open(file)
                    img = img.resize((1280, 720), Image.Resampling.LANCZOS)
                    img.save(filepath)

                    portada = f'uploads/productos/{filename}'

            # Create product
            producto = Producto(
                titulo=titulo,
                ruta=ruta,
                id_categoria=id_categoria,
                id_subcategoria=id_subcategoria,
                tipo=tipo,
                precio=precio,
                stock=stock,
                descripcion=descripcion,
                titular=titular,
                portada=portada,
                estado=1
            )

            db.session.add(producto)
            db.session.commit()

            flash('Producto creado exitosamente!', 'success')
            return redirect(url_for('admin.products'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear producto: {e}', 'error')

    categorias = Categoria.query.all()
    subcategorias = Subcategoria.query.all()
    return render_template('admin/product_create.html', categorias=categorias, subcategorias=subcategorias)


@admin_bp.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_product(id):
    """Edit product."""
    producto = Producto.query.get_or_404(id)

    if request.method == 'POST':
        try:
            producto.titulo = request.form.get('titulo')
            producto.ruta = request.form.get('ruta')
            producto.id_categoria = request.form.get('id_categoria', type=int)
            producto.id_subcategoria = request.form.get('id_subcategoria', type=int) or None
            producto.tipo = request.form.get('tipo', 'fisico')
            producto.precio = float(request.form.get('precio', 0))
            producto.stock = int(request.form.get('stock', 0)) if producto.tipo == 'fisico' else 0
            producto.stock_minimo = int(request.form.get('stock_minimo', 5))
            producto.descripcion = request.form.get('descripcion', '')
            producto.titular = request.form.get('titular', '')

            # Handle offer
            producto.oferta = 1 if request.form.get('oferta') == '1' else 0
            if producto.oferta:
                producto.precioOferta = float(request.form.get('precioOferta', 0))
                producto.descuentoOferta = int(request.form.get('descuentoOferta', 0))
                fin_oferta = request.form.get('finOferta')
                if fin_oferta:
                    producto.finOferta = datetime.strptime(fin_oferta, '%Y-%m-%d')

            # Handle image upload
            if 'portada' in request.files:
                file = request.files['portada']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    upload_folder = os.path.join('app/static/uploads/productos')
                    if not os.path.exists(upload_folder):
                        os.makedirs(upload_folder)

                    filepath = os.path.join(upload_folder, filename)

                    # Resize image to 1280x720
                    img = Image.open(file)
                    img = img.resize((1280, 720), Image.Resampling.LANCZOS)
                    img.save(filepath)

                    producto.portada = f'uploads/productos/{filename}'

            db.session.commit()
            flash('Producto actualizado exitosamente!', 'success')
            return redirect(url_for('admin.products'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar producto: {e}', 'error')

    categorias = Categoria.query.all()
    subcategorias = Subcategoria.query.all()
    return render_template('admin/product_edit.html',
                         producto=producto,
                         categorias=categorias,
                         subcategorias=subcategorias)


@admin_bp.route('/products/delete/<int:id>', methods=['POST'])
@admin_required
def delete_product(id):
    """Delete product."""
    try:
        producto = Producto.query.get_or_404(id)

        # Check if product has orders
        if producto.compras.count() > 0:
            flash(f'No se puede eliminar "{producto.titulo}" porque tiene {producto.compras.count()} compra(s) asociada(s). Desactívelo en su lugar.', 'error')
            return redirect(url_for('admin.products'))

        # Check if product has comments
        if producto.comentarios.count() > 0:
            flash(f'No se puede eliminar "{producto.titulo}" porque tiene {producto.comentarios.count()} comentario(s). Desactívelo en su lugar.', 'error')
            return redirect(url_for('admin.products'))

        # Check if product is in wishlists
        if producto.deseos.count() > 0:
            flash(f'No se puede eliminar "{producto.titulo}" porque está en {producto.deseos.count()} lista(s) de deseos. Desactívelo en su lugar.', 'error')
            return redirect(url_for('admin.products'))

        # Safe to delete
        db.session.delete(producto)
        db.session.commit()
        flash(f'Producto "{producto.titulo}" eliminado exitosamente!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar producto: {e}', 'error')

    return redirect(url_for('admin.products'))


@admin_bp.route('/products/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_product(id):
    """Toggle product status."""
    try:
        producto = Producto.query.get_or_404(id)
        producto.estado = 0 if producto.estado == 1 else 1
        db.session.commit()
        return jsonify({'success': True, 'estado': producto.estado})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/products/<int:id>/gallery', methods=['GET', 'POST'])
@admin_required
def product_gallery(id):
    """Manage product image gallery."""
    producto = Producto.query.get_or_404(id)

    if request.method == 'POST':
        # Handle multiple image uploads
        uploaded_files = request.files.getlist('images')

        if not uploaded_files:
            flash('No se seleccionaron imágenes.', 'error')
            return redirect(url_for('admin.product_gallery', id=id))

        upload_folder = os.path.join('app/static/uploads/productos')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        uploaded_count = 0
        for file in uploaded_files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                # Add timestamp to avoid name conflicts
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                filename = f"{timestamp}_{filename}"
                filepath = os.path.join(upload_folder, filename)

                # Resize image to 1280x720
                try:
                    img = Image.open(file)
                    img = img.resize((1280, 720), Image.Resampling.LANCZOS)
                    img.save(filepath)

                    # Add to multimedia gallery
                    image_path = f'uploads/productos/{filename}'
                    producto.add_multimedia_image(image_path)
                    uploaded_count += 1
                except Exception as e:
                    flash(f'Error al procesar imagen {file.filename}: {e}', 'error')

        if uploaded_count > 0:
            flash(f'{uploaded_count} imagen(es) agregada(s) exitosamente.', 'success')

        return redirect(url_for('admin.product_gallery', id=id))

    # GET request - show gallery management page
    return render_template('admin/product_gallery.html', producto=producto)


@admin_bp.route('/products/<int:id>/gallery/delete/<path:image>', methods=['POST'])
@admin_required
def delete_gallery_image(id, image):
    """Delete image from product gallery."""
    try:
        producto = Producto.query.get_or_404(id)

        # Remove from multimedia list
        producto.remove_multimedia_image(image)

        # Delete physical file
        try:
            img_path = os.path.join('app/static', image)
            if os.path.exists(img_path):
                os.remove(img_path)
        except Exception as e:
            print(f'Error deleting image file: {e}')

        return jsonify({'success': True, 'message': 'Imagen eliminada exitosamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@admin_bp.route('/orders')
@admin_required
def orders():
    """Manage orders."""
    # Get statistics for cards
    from sqlalchemy import func
    total_orders = Compra.query.count()
    total_revenue = db.session.query(func.sum(Compra.pago)).scalar() or 0
    pending_orders = Compra.query.filter_by(estado='pendiente').count()
    completed_orders = Compra.query.filter_by(estado='completado').count()

    return render_template('admin/orders.html',
                         total_orders=total_orders,
                         total_revenue=total_revenue,
                         pending_orders=pending_orders,
                         completed_orders=completed_orders)


@admin_bp.route('/orders/ajax')
@admin_required
def orders_ajax():
    """AJAX endpoint for DataTables order listing."""
    # DataTables parameters
    draw = request.args.get('draw', type=int, default=1)
    start = request.args.get('start', type=int, default=0)
    length = request.args.get('length', type=int, default=10)
    search_value = request.args.get('search[value]', default='')
    order_column = request.args.get('order[0][column]', type=int, default=0)
    order_dir = request.args.get('order[0][dir]', default='desc')

    # Custom filters
    metodo_filter = request.args.get('metodo_filter', default='')
    estado_filter = request.args.get('estado_filter', default='')
    fecha_desde = request.args.get('fecha_desde', default='')
    fecha_hasta = request.args.get('fecha_hasta', default='')
    usuario_filter = request.args.get('usuario_filter', default='')

    # Base query with joins
    query = Compra.query.join(User, Compra.id_usuario == User.id, isouter=True)\
                        .join(Producto, Compra.id_producto == Producto.id, isouter=True)

    # Search filter
    if search_value:
        query = query.filter(
            db.or_(
                User.nombre.like(f'%{search_value}%'),
                User.email.like(f'%{search_value}%'),
                Producto.titulo.like(f'%{search_value}%'),
                Compra.email.like(f'%{search_value}%'),
                Compra.tracking.like(f'%{search_value}%') if hasattr(Compra, 'tracking') else False
            )
        )

    # Method filter
    if metodo_filter:
        query = query.filter(Compra.metodo == metodo_filter)

    # Status filter
    if estado_filter:
        query = query.filter(Compra.estado == estado_filter)

    # Date range filter
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Compra.fecha >= fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            # Add 1 day to include the entire end date
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            from datetime import timedelta
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            query = query.filter(Compra.fecha < fecha_hasta_dt)
        except ValueError:
            pass

    # User filter
    if usuario_filter:
        query = query.filter(
            db.or_(
                User.nombre.like(f'%{usuario_filter}%'),
                User.email.like(f'%{usuario_filter}%')
            )
        )

    # Total records
    total_records = Compra.query.count()
    filtered_records = query.count()

    # Ordering
    columns = [
        Compra.id,
        User.nombre,
        Producto.titulo,
        Compra.cantidad,
        Compra.pago,
        Compra.metodo,
        Compra.estado,
        Compra.fecha
    ]

    if 0 <= order_column < len(columns):
        order_col = columns[order_column]
        if order_dir == 'asc':
            query = query.order_by(order_col.asc())
        else:
            query = query.order_by(order_col.desc())
    else:
        query = query.order_by(Compra.fecha.desc())

    # Get paginated data
    orders = query.offset(start).limit(length).all()

    # Format data for DataTables
    data = []
    for order in orders:
        # Cliente column
        if order.usuario:
            cliente_html = f'<strong>{order.usuario.nombre}</strong><br><small class="text-muted">{order.usuario.email}</small>'
        else:
            cliente_html = f'<small class="text-muted">{order.email or "N/A"}</small>'

        # Producto column
        if order.producto:
            producto_html = f'{order.producto.titulo[:40]}{"..." if len(order.producto.titulo) > 40 else ""}'
        else:
            producto_html = 'N/A'

        # Método de pago
        metodo_icons = {
            'paypal': '<i class="fab fa-paypal text-primary"></i> PayPal',
            'paymentez': '<i class="fas fa-credit-card text-success"></i> Paymentez',
            'datafast': '<i class="fas fa-credit-card text-info"></i> Datafast',
            'deuna': '<i class="fas fa-mobile-alt text-warning"></i> De Una',
            'transferencia': '<i class="fas fa-university text-secondary"></i> Transferencia',
            'transferencia_comprobante': '<i class="fas fa-university text-secondary"></i> Transferencia'
        }
        metodo_html = metodo_icons.get(order.metodo, order.metodo)

        # Estado badge
        estado_badges = {
            'completado': 'bg-success',
            'pendiente': 'bg-warning',
            'procesando': 'bg-info',
            'enviado': 'bg-primary',
            'entregado': 'bg-success',
            'cancelado': 'bg-danger'
        }
        estado_class = estado_badges.get(order.estado, 'bg-secondary')
        estado_html = f'<span class="badge {estado_class}">{order.estado.capitalize()}</span>'

        data.append({
            'id': order.id,
            'cliente': cliente_html,
            'producto': producto_html,
            'cantidad': order.cantidad,
            'total': f'${order.pago:.2f}',
            'metodo': metodo_html,
            'estado': order.estado,
            'estado_html': estado_html,
            'fecha': order.fecha.strftime('%d/%m/%Y %H:%M') if order.fecha else 'N/A',
            'tracking': order.tracking if hasattr(order, 'tracking') else ''
        })

    return jsonify({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@admin_bp.route('/orders/update-status/<int:id>', methods=['POST'])
@admin_required
def update_order_status(id):
    """Update order status."""
    try:
        order = Compra.query.get_or_404(id)
        estado = request.form.get('estado')
        tracking = request.form.get('tracking', '')
        
        if estado:
            order.estado = estado
            if tracking:
                order.tracking = tracking
            order.fecha_estado = datetime.now()
            db.session.commit()
            
            flash('Estado de orden actualizado correctamente!', 'success')
        else:
            flash('Debe seleccionar un estado.', 'error')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar estado: {e}', 'error')
    
    return redirect(url_for('admin.orders'))



@admin_bp.route('/analytics')
@admin_required
def analytics():
    """Analytics and reports."""
    # Get visit statistics
    visits_by_country = VisitaPais.query.order_by(VisitaPais.cantidad.desc()).limit(10).all()
    total_visits = VisitaPersona.get_total_visits()
    unique_visitors = VisitaPersona.get_unique_visitors()

    return render_template('admin/analytics.html',
                         visits_by_country=visits_by_country,
                         total_visits=total_visits,
                         unique_visitors=unique_visitors)


# ===========================
# ADVANCED REPORTS
# ===========================

@admin_bp.route('/reports')
@admin_required
def reports():
    """Advanced sales reports with charts."""
    from sqlalchemy import func
    from datetime import datetime, timedelta

    # Default date range: last 30 days
    fecha_hasta = datetime.now()
    fecha_desde = fecha_hasta - timedelta(days=30)

    # Get all products for filter
    productos = Producto.query.order_by(Producto.titulo).all()

    # Get all users for filter
    usuarios = User.query.order_by(User.nombre).all()

    # Initial statistics
    total_ventas = Compra.query.filter(Compra.estado != 'cancelado').count()
    ingresos_totales = db.session.query(func.sum(Compra.pago)).filter(Compra.estado != 'cancelado').scalar() or 0
    ticket_promedio = ingresos_totales / total_ventas if total_ventas > 0 else 0

    return render_template('admin/reports.html',
                         productos=productos,
                         usuarios=usuarios,
                         total_ventas=total_ventas,
                         ingresos_totales=ingresos_totales,
                         ticket_promedio=ticket_promedio)


@admin_bp.route('/reports/data')
@admin_required
def reports_data():
    """AJAX endpoint for chart data."""
    from sqlalchemy import func
    from datetime import datetime, timedelta

    # Get filter parameters
    fecha_desde = request.args.get('fecha_desde', default='')
    fecha_hasta = request.args.get('fecha_hasta', default='')
    producto_id = request.args.get('producto_id', default='')
    usuario_id = request.args.get('usuario_id', default='')
    metodo = request.args.get('metodo', default='')

    # Build query
    query = Compra.query.filter(Compra.estado != 'cancelado')

    # Apply filters
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Compra.fecha >= fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            query = query.filter(Compra.fecha < fecha_hasta_dt)
        except ValueError:
            pass

    if producto_id:
        query = query.filter(Compra.id_producto == int(producto_id))

    if usuario_id:
        query = query.filter(Compra.id_usuario == int(usuario_id))

    if metodo:
        query = query.filter(Compra.metodo == metodo)

    # Sales by date
    ventas_por_fecha = db.session.query(
        func.date(Compra.fecha).label('fecha'),
        func.count(Compra.id).label('cantidad'),
        func.sum(Compra.pago).label('total')
    ).filter(Compra.estado != 'cancelado')

    # Apply same filters
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            ventas_por_fecha = ventas_por_fecha.filter(Compra.fecha >= fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            ventas_por_fecha = ventas_por_fecha.filter(Compra.fecha < fecha_hasta_dt)
        except ValueError:
            pass

    if producto_id:
        ventas_por_fecha = ventas_por_fecha.filter(Compra.id_producto == int(producto_id))

    if usuario_id:
        ventas_por_fecha = ventas_por_fecha.filter(Compra.id_usuario == int(usuario_id))

    if metodo:
        ventas_por_fecha = ventas_por_fecha.filter(Compra.metodo == metodo)

    ventas_por_fecha = ventas_por_fecha.group_by(func.date(Compra.fecha)).order_by(func.date(Compra.fecha)).all()

    # Top products
    top_productos = db.session.query(
        Producto.titulo,
        func.count(Compra.id).label('ventas'),
        func.sum(Compra.pago).label('ingresos')
    ).join(Compra, Compra.id_producto == Producto.id)\
     .filter(Compra.estado != 'cancelado')

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            top_productos = top_productos.filter(Compra.fecha >= fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            top_productos = top_productos.filter(Compra.fecha < fecha_hasta_dt)
        except ValueError:
            pass

    top_productos = top_productos.group_by(Producto.titulo)\
                                 .order_by(func.sum(Compra.pago).desc())\
                                 .limit(10).all()

    # Sales by payment method
    ventas_por_metodo = db.session.query(
        Compra.metodo,
        func.count(Compra.id).label('cantidad'),
        func.sum(Compra.pago).label('total')
    ).filter(Compra.estado != 'cancelado')

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            ventas_por_metodo = ventas_por_metodo.filter(Compra.fecha >= fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            ventas_por_metodo = ventas_por_metodo.filter(Compra.fecha < fecha_hasta_dt)
        except ValueError:
            pass

    ventas_por_metodo = ventas_por_metodo.group_by(Compra.metodo).all()

    # Format data for Chart.js
    return jsonify({
        'ventas_por_fecha': {
            'labels': [str(v.fecha) for v in ventas_por_fecha],
            'cantidad': [v.cantidad for v in ventas_por_fecha],
            'ingresos': [float(v.total or 0) for v in ventas_por_fecha]
        },
        'top_productos': {
            'labels': [p.titulo[:30] for p in top_productos],
            'ventas': [p.ventas for p in top_productos],
            'ingresos': [float(p.ingresos or 0) for p in top_productos]
        },
        'ventas_por_metodo': {
            'labels': [v.metodo for v in ventas_por_metodo],
            'cantidad': [v.cantidad for v in ventas_por_metodo],
            'total': [float(v.total or 0) for v in ventas_por_metodo]
        },
        'estadisticas': {
            'total_ventas': query.count(),
            'ingresos_totales': float(db.session.query(func.sum(Compra.pago)).filter(Compra.id.in_([c.id for c in query.all()])).scalar() or 0)
        }
    })


@admin_bp.route('/reports/export')
@admin_required
def export_reports():
    """Export sales report to Excel."""
    from sqlalchemy import func
    from datetime import datetime, timedelta
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        flash('openpyxl no está instalado. Ejecuta: pip install openpyxl', 'error')
        return redirect(url_for('admin.reports'))

    # Get filter parameters
    fecha_desde = request.args.get('fecha_desde', default='')
    fecha_hasta = request.args.get('fecha_hasta', default='')
    producto_id = request.args.get('producto_id', default='')
    usuario_id = request.args.get('usuario_id', default='')
    metodo = request.args.get('metodo', default='')

    # Build query
    query = Compra.query.filter(Compra.estado != 'cancelado')\
                        .join(User, Compra.id_usuario == User.id, isouter=True)\
                        .join(Producto, Compra.id_producto == Producto.id, isouter=True)

    # Apply filters
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Compra.fecha >= fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt + timedelta(days=1)
            query = query.filter(Compra.fecha < fecha_hasta_dt)
        except ValueError:
            pass

    if producto_id:
        query = query.filter(Compra.id_producto == int(producto_id))

    if usuario_id:
        query = query.filter(Compra.id_usuario == int(usuario_id))

    if metodo:
        query = query.filter(Compra.metodo == metodo)

    ventas = query.order_by(Compra.fecha.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Add headers
    headers = ['ID', 'Fecha', 'Cliente', 'Email', 'Producto', 'Cantidad', 'Total', 'Método Pago', 'Estado']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add data
    for row_num, venta in enumerate(ventas, 2):
        ws.cell(row=row_num, column=1, value=venta.id)
        ws.cell(row=row_num, column=2, value=venta.fecha.strftime('%Y-%m-%d %H:%M') if venta.fecha else '')
        ws.cell(row=row_num, column=3, value=venta.usuario.nombre if venta.usuario else venta.email or 'N/A')
        ws.cell(row=row_num, column=4, value=venta.usuario.email if venta.usuario else venta.email or 'N/A')
        ws.cell(row=row_num, column=5, value=venta.producto.titulo if venta.producto else 'N/A')
        ws.cell(row=row_num, column=6, value=venta.cantidad)
        ws.cell(row=row_num, column=7, value=float(venta.pago))
        ws.cell(row=row_num, column=8, value=venta.metodo)
        ws.cell(row=row_num, column=9, value=venta.estado)

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 12

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    filename = f'reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ===========================
# SEO HEADERS MANAGEMENT
# ===========================

@admin_bp.route('/seo-headers')
@admin_required
def seo_headers():
    """List all SEO headers."""
    cabeceras = Cabecera.query.order_by(Cabecera.fecha.desc()).all()
    total_cabeceras = Cabecera.query.count()

    return render_template('admin/seo_headers.html',
                         cabeceras=cabeceras,
                         total_cabeceras=total_cabeceras)


@admin_bp.route('/seo-headers/create', methods=['GET', 'POST'])
@admin_required
def create_seo_header():
    """Create new SEO header."""
    if request.method == 'POST':
        try:
            ruta = request.form.get('ruta', '').strip()
            titulo = request.form.get('titulo', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            palabras_claves = request.form.get('palabras_claves', '').strip()

            # Validate required fields
            if not ruta or not titulo:
                flash('La ruta y el título son obligatorios.', 'error')
                return render_template('admin/seo_header_form.html')

            # Check if route already exists
            existing = Cabecera.query.filter_by(ruta=ruta).first()
            if existing:
                flash(f'Ya existe una cabecera SEO para la ruta "{ruta}".', 'error')
                return render_template('admin/seo_header_form.html')

            # Handle image upload
            portada_path = ''
            if 'portada' in request.files:
                file = request.files['portada']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"seo_{timestamp}_{filename}"

                    upload_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'app/static/uploads'), 'seo')
                    os.makedirs(upload_folder, exist_ok=True)
                    filepath = os.path.join(upload_folder, filename)

                    # Save and resize image
                    from PIL import Image
                    img = Image.open(file)
                    img.thumbnail((1200, 630))  # Open Graph recommended size
                    img.save(filepath, optimize=True, quality=85)

                    portada_path = f"uploads/seo/{filename}"

            # Create cabecera
            cabecera = Cabecera(
                ruta=ruta,
                titulo=titulo,
                descripcion=descripcion,
                palabrasClaves=palabras_claves,
                portada=portada_path
            )

            db.session.add(cabecera)
            db.session.commit()

            flash(f'Cabecera SEO para "{ruta}" creada exitosamente.', 'success')
            return redirect(url_for('admin.seo_headers'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear cabecera SEO: {e}', 'error')

    return render_template('admin/seo_header_form.html')


@admin_bp.route('/seo-headers/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_seo_header(id):
    """Edit SEO header."""
    cabecera = Cabecera.query.get_or_404(id)

    if request.method == 'POST':
        try:
            ruta = request.form.get('ruta', '').strip()
            titulo = request.form.get('titulo', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            palabras_claves = request.form.get('palabras_claves', '').strip()

            # Validate required fields
            if not ruta or not titulo:
                flash('La ruta y el título son obligatorios.', 'error')
                return render_template('admin/seo_header_form.html', cabecera=cabecera)

            # Check if route already exists (excluding current)
            existing = Cabecera.query.filter(Cabecera.ruta == ruta, Cabecera.id != id).first()
            if existing:
                flash(f'Ya existe otra cabecera SEO para la ruta "{ruta}".', 'error')
                return render_template('admin/seo_header_form.html', cabecera=cabecera)

            # Update fields
            cabecera.ruta = ruta
            cabecera.titulo = titulo
            cabecera.descripcion = descripcion
            cabecera.palabrasClaves = palabras_claves

            # Handle image upload
            if 'portada' in request.files:
                file = request.files['portada']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"seo_{timestamp}_{filename}"

                    upload_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'app/static/uploads'), 'seo')
                    os.makedirs(upload_folder, exist_ok=True)
                    filepath = os.path.join(upload_folder, filename)

                    # Save and resize image
                    from PIL import Image
                    img = Image.open(file)
                    img.thumbnail((1200, 630))
                    img.save(filepath, optimize=True, quality=85)

                    cabecera.portada = f"uploads/seo/{filename}"

            db.session.commit()

            flash(f'Cabecera SEO para "{ruta}" actualizada exitosamente.', 'success')
            return redirect(url_for('admin.seo_headers'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar cabecera SEO: {e}', 'error')

    return render_template('admin/seo_header_form.html', cabecera=cabecera)


@admin_bp.route('/seo-headers/delete/<int:id>', methods=['POST'])
@admin_required
def delete_seo_header(id):
    """Delete SEO header."""
    try:
        cabecera = Cabecera.query.get_or_404(id)
        ruta = cabecera.ruta

        db.session.delete(cabecera)
        db.session.commit()

        flash(f'Cabecera SEO para "{ruta}" eliminada exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar cabecera SEO: {e}', 'error')

    return redirect(url_for('admin.seo_headers'))


@admin_bp.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings():
    """Store settings and payment gateways configuration."""
    from app.models.comercio import Comercio
    import json

    # Auto-migrate: Add SMTP columns if they don't exist
    try:
        migrations = [
            "ALTER TABLE comercio ADD COLUMN IF NOT EXISTS mailServer VARCHAR(100) DEFAULT 'smtp.gmail.com'",
            "ALTER TABLE comercio ADD COLUMN IF NOT EXISTS mailPort INT DEFAULT 587",
            "ALTER TABLE comercio ADD COLUMN IF NOT EXISTS mailUseTLS BOOLEAN DEFAULT TRUE",
            "ALTER TABLE comercio ADD COLUMN IF NOT EXISTS mailUsername VARCHAR(255)",
            "ALTER TABLE comercio ADD COLUMN IF NOT EXISTS mailPassword TEXT",
            "ALTER TABLE comercio ADD COLUMN IF NOT EXISTS mailDefaultSender VARCHAR(255)"
        ]
        for migration in migrations:
            try:
                db.session.execute(db.text(migration))
            except Exception:
                pass  # Column already exists
        db.session.commit()
    except Exception:
        db.session.rollback()

    config = Comercio.get_config()

    if request.method == 'POST':
        try:
            # Store settings
            config.impuesto = float(request.form.get('impuesto', 0))
            config.envioNacional = float(request.form.get('envioNacional', 0))
            config.envioInternacional = float(request.form.get('envioInternacional', 0))
            config.pais = request.form.get('pais', 'Ecuador')

            # PayPal settings
            config.modoPaypal = request.form.get('modoPaypal', 'sandbox')
            config.clienteIdPaypal = request.form.get('clienteIdPaypal', '')
            config.llaveSecretaPaypal = request.form.get('llaveSecretaPaypal', '')

            # Paymentez settings
            config.modoPaymentez = request.form.get('modoPaymentez', 'test')
            config.appCodePaymentez = request.form.get('appCodePaymentez', '')
            config.appKeyPaymentez = request.form.get('appKeyPaymentez', '')

            # Datafast settings
            config.modoDatafast = request.form.get('modoDatafast', 'test')
            config.midDatafast = request.form.get('midDatafast', '')
            config.tidDatafast = request.form.get('tidDatafast', '')

            # De Una settings
            config.modoDeUna = request.form.get('modoDeUna', 'test')
            config.apiKeyDeUna = request.form.get('apiKeyDeUna', '')

            # SMTP Email settings
            config.mailServer = request.form.get('mailServer', 'smtp.gmail.com')
            config.mailPort = int(request.form.get('mailPort', 587))
            config.mailUseTLS = request.form.get('mailUseTLS') == 'true'
            config.mailUsername = request.form.get('mailUsername', '')
            config.mailPassword = request.form.get('mailPassword', '')
            config.mailDefaultSender = request.form.get('mailDefaultSender', '')

            # Bank accounts (save as JSON)
            bank_accounts = {}
            for bank in ['banco_pichincha', 'banco_guayaquil', 'banco_pacifico']:
                bank_accounts[bank] = {
                    'nombre': request.form.get(f'{bank}_nombre', ''),
                    'cuenta': request.form.get(f'{bank}_cuenta', ''),
                    'tipo': request.form.get(f'{bank}_tipo', 'Ahorros'),
                    'cedula': request.form.get(f'{bank}_cedula', '')
                }
            config.cuentasBancarias = json.dumps(bank_accounts)

            db.session.commit()
            flash('Configuración guardada exitosamente!', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar configuración: {e}', 'error')

        return redirect(url_for('admin.settings'))

    # Get bank accounts
    bank_accounts = config.get_bank_accounts()
    if not bank_accounts:
        bank_accounts = {
            'banco_pichincha': {'nombre': '', 'cuenta': '', 'tipo': 'Ahorros', 'cedula': ''},
            'banco_guayaquil': {'nombre': '', 'cuenta': '', 'tipo': 'Ahorros', 'cedula': ''},
            'banco_pacifico': {'nombre': '', 'cuenta': '', 'tipo': 'Ahorros', 'cedula': ''}
        }

    return render_template('admin/settings.html',
                         config=config,
                         bank_accounts=bank_accounts)


@admin_bp.route('/export/users')
@admin_required
def export_users():
    """Export users to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Headers
        headers = ['ID', 'Nombre', 'Email', 'Modo', 'Verificado', 'Fecha Registro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        
        # Data
        users = User.query.all()
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.nombre)
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.modo)
            ws.cell(row=row, column=5, value='Sí' if user.verificacion == 0 else 'No')
            ws.cell(row=row, column=6, value=user.fecha.strftime('%d/%m/%Y'))
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'usuarios_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        flash(f'Error al exportar: {e}', 'error')
        return redirect(url_for('admin.users'))


@admin_bp.route('/export/products')
@admin_required
def export_products():
    """Export products to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Headers
        headers = ['ID', 'Título', 'Categoría', 'Precio', 'Stock', 'Ventas', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        
        # Data
        products = Producto.query.all()
        for row, p in enumerate(products, 2):
            ws.cell(row=row, column=1, value=p.id)
            ws.cell(row=row, column=2, value=p.titulo)
            ws.cell(row=row, column=3, value=p.categoria.categoria if p.categoria else 'N/A')
            ws.cell(row=row, column=4, value=float(p.precio))
            ws.cell(row=row, column=5, value=p.stock)
            ws.cell(row=row, column=6, value=p.ventas)
            ws.cell(row=row, column=7, value='Activo' if p.estado == 1 else 'Inactivo')
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'productos_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        flash(f'Error al exportar: {e}', 'error')
        return redirect(url_for('admin.products'))


@admin_bp.route('/export/orders')
@admin_required
def export_orders():
    """Export orders to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        
        # Headers
        headers = ['ID', 'Cliente', 'Producto', 'Cantidad', 'Total', 'Método', 'Estado', 'Fecha']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        
        # Data
        orders = Compra.query.all()
        for row, o in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=o.id)
            ws.cell(row=row, column=2, value=o.usuario.nombre if o.usuario else 'N/A')
            ws.cell(row=row, column=3, value=o.producto.titulo if o.producto else 'N/A')
            ws.cell(row=row, column=4, value=o.cantidad)
            ws.cell(row=row, column=5, value=float(o.pago))
            ws.cell(row=row, column=6, value=o.metodo)
            ws.cell(row=row, column=7, value=o.estado)
            ws.cell(row=row, column=8, value=o.fecha.strftime('%d/%m/%Y %H:%M'))
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'pedidos_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        flash(f'Error al exportar: {e}', 'error')
        return redirect(url_for('admin.orders'))


# ===========================
# CATEGORY MANAGEMENT
# ===========================

@admin_bp.route('/categories')
@admin_required
def categories():
    """Categories management page."""
    # Get statistics
    total_categories = Categoria.query.count()
    active_categories = Categoria.query.filter_by(estado=1).count()

    return render_template('admin/categories.html',
                         total_categories=total_categories,
                         active_categories=active_categories)


@admin_bp.route('/categories/ajax')
@admin_required
def categories_ajax():
    """AJAX endpoint for DataTables category listing."""
    # DataTables parameters
    draw = request.args.get('draw', type=int, default=1)
    start = request.args.get('start', type=int, default=0)
    length = request.args.get('length', type=int, default=10)
    search_value = request.args.get('search[value]', default='')
    order_column = request.args.get('order[0][column]', type=int, default=0)
    order_dir = request.args.get('order[0][dir]', default='desc')

    # Base query
    query = Categoria.query

    # Search filter
    if search_value:
        query = query.filter(
            db.or_(
                Categoria.categoria.like(f'%{search_value}%'),
                Categoria.ruta.like(f'%{search_value}%')
            )
        )

    # Total records
    total_records = Categoria.query.count()
    filtered_records = query.count()

    # Ordering
    columns = [
        Categoria.id,
        Categoria.categoria,
        Categoria.ruta,
        None,  # productos count (not sortable)
        Categoria.estado,
        Categoria.fecha
    ]

    if 0 <= order_column < len(columns) and columns[order_column] is not None:
        order_col = columns[order_column]
        if order_dir == 'asc':
            query = query.order_by(order_col.asc())
        else:
            query = query.order_by(order_col.desc())
    else:
        query = query.order_by(Categoria.fecha.desc())

    # Get paginated data
    categorias = query.offset(start).limit(length).all()

    # Format data for DataTables
    data = []
    for categoria in categorias:
        data.append({
            'id': categoria.id,
            'categoria': categoria.categoria,
            'ruta': f'<code>{categoria.ruta}</code>',
            'productos': categoria.get_products_count(),
            'estado': categoria.estado,
            'fecha': categoria.fecha.strftime('%d/%m/%Y') if categoria.fecha else 'N/A'
        })

    return jsonify({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@admin_bp.route('/categories/create', methods=['GET', 'POST'])
@admin_required
def create_category():
    """Create new category."""
    if request.method == 'POST':
        categoria_nombre = request.form.get('categoria')
        ruta = request.form.get('ruta')
        estado = request.form.get('estado', 1, type=int)
        
        if not categoria_nombre or not ruta:
            flash('Todos los campos son requeridos.', 'error')
            return redirect(url_for('admin.create_category'))
        
        # Check if ruta already exists
        existing = Categoria.query.filter_by(ruta=ruta).first()
        if existing:
            flash('La ruta ya existe. Use una ruta única.', 'error')
            return redirect(url_for('admin.create_category'))
        
        categoria = Categoria(
            categoria=categoria_nombre,
            ruta=ruta,
            estado=estado
        )
        
        db.session.add(categoria)
        db.session.commit()
        
        flash('Categoría creada exitosamente.', 'success')
        return redirect(url_for('admin.categories'))
    
    return render_template('admin/category_create.html')


@admin_bp.route('/categories/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_category(id):
    """Edit category."""
    categoria = Categoria.query.get_or_404(id)
    
    if request.method == 'POST':
        categoria_nombre = request.form.get('categoria')
        ruta = request.form.get('ruta')
        estado = request.form.get('estado', 1, type=int)
        
        if not categoria_nombre or not ruta:
            flash('Todos los campos son requeridos.', 'error')
            return redirect(url_for('admin.edit_category', id=id))
        
        # Check if ruta already exists (except current)
        existing = Categoria.query.filter(Categoria.ruta == ruta, Categoria.id != id).first()
        if existing:
            flash('La ruta ya existe. Use una ruta única.', 'error')
            return redirect(url_for('admin.edit_category', id=id))
        
        categoria.categoria = categoria_nombre
        categoria.ruta = ruta
        categoria.estado = estado
        
        db.session.commit()
        
        flash('Categoría actualizada exitosamente.', 'success')
        return redirect(url_for('admin.categories'))
    
    return render_template('admin/category_edit.html', categoria=categoria)


@admin_bp.route('/categories/delete/<int:id>', methods=['POST'])
@admin_required
def delete_category(id):
    """Delete category."""
    categoria = Categoria.query.get_or_404(id)
    
    # Check if category has products
    if categoria.get_products_count() > 0:
        flash('No se puede eliminar una categoría con productos asignados.', 'error')
        return redirect(url_for('admin.categories'))
    
    db.session.delete(categoria)
    db.session.commit()
    
    flash('Categoría eliminada exitosamente.', 'success')
    return redirect(url_for('admin.categories'))


@admin_bp.route('/categories/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_category(id):
    """Toggle category status."""
    categoria = Categoria.query.get_or_404(id)
    categoria.estado = 0 if categoria.estado == 1 else 1
    db.session.commit()

    return jsonify({'success': True, 'estado': categoria.estado})


# ===========================
# SUBCATEGORIES MANAGEMENT
# ===========================

@admin_bp.route('/subcategories')
@admin_required
def subcategories():
    """Subcategories management page."""
    # Get statistics
    total_subcategories = Subcategoria.query.count()
    active_subcategories = Subcategoria.query.filter_by(estado=1).count()
    categorias = Categoria.query.all()

    return render_template('admin/subcategories.html',
                          total_subcategories=total_subcategories,
                          active_subcategories=active_subcategories,
                          categorias=categorias)


@admin_bp.route('/subcategories/ajax')
@admin_required
def subcategories_ajax():
    """AJAX endpoint for DataTables subcategory listing."""
    # DataTables parameters
    draw = request.args.get('draw', type=int, default=1)
    start = request.args.get('start', type=int, default=0)
    length = request.args.get('length', type=int, default=10)
    search_value = request.args.get('search[value]', default='')
    order_column = request.args.get('order[0][column]', type=int, default=0)
    order_dir = request.args.get('order[0][dir]', default='desc')

    # Custom filter
    categoria_filter = request.args.get('categoria_filter', default='')

    # Base query with join
    query = Subcategoria.query.join(Categoria, Subcategoria.id_categoria == Categoria.id, isouter=True)

    # Search filter
    if search_value:
        query = query.filter(
            db.or_(
                Subcategoria.subcategoria.like(f'%{search_value}%'),
                Subcategoria.ruta.like(f'%{search_value}%'),
                Categoria.categoria.like(f'%{search_value}%')
            )
        )

    # Category filter
    if categoria_filter:
        query = query.filter(Subcategoria.id_categoria == int(categoria_filter))

    # Total records
    total_records = Subcategoria.query.count()
    filtered_records = query.count()

    # Ordering
    columns = [
        Subcategoria.id,
        Subcategoria.subcategoria,
        Categoria.categoria,
        Subcategoria.ruta,
        None,  # productos count (not sortable)
        Subcategoria.estado,
        Subcategoria.fecha
    ]

    if 0 <= order_column < len(columns) and columns[order_column] is not None:
        order_col = columns[order_column]
        if order_dir == 'asc':
            query = query.order_by(order_col.asc())
        else:
            query = query.order_by(order_col.desc())
    else:
        query = query.order_by(Subcategoria.fecha.desc())

    # Get paginated data
    subcategorias = query.offset(start).limit(length).all()

    # Format data for DataTables
    data = []
    for subcat in subcategorias:
        data.append({
            'id': subcat.id,
            'subcategoria': subcat.subcategoria,
            'categoria': subcat.categoria.categoria if subcat.categoria else 'N/A',
            'categoria_id': subcat.id_categoria,
            'ruta': f'<code>{subcat.ruta}</code>',
            'productos': subcat.get_products_count(),
            'estado': subcat.estado,
            'fecha': subcat.fecha.strftime('%d/%m/%Y') if subcat.fecha else 'N/A'
        })

    return jsonify({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@admin_bp.route('/subcategories/create', methods=['GET', 'POST'])
@admin_required
def create_subcategory():
    """Create new subcategory."""
    if request.method == 'POST':
        subcategoria_nombre = request.form.get('subcategoria')
        ruta = request.form.get('ruta')
        id_categoria = request.form.get('id_categoria', type=int)
        estado = request.form.get('estado', 1, type=int)

        if not subcategoria_nombre or not ruta or not id_categoria:
            flash('Todos los campos son requeridos.', 'error')
            return redirect(url_for('admin.create_subcategory'))

        # Check if ruta already exists
        existing = Subcategoria.query.filter_by(ruta=ruta).first()
        if existing:
            flash('La ruta ya existe. Use una ruta única.', 'error')
            return redirect(url_for('admin.create_subcategory'))

        # Verify parent category exists
        categoria = Categoria.query.get(id_categoria)
        if not categoria:
            flash('Categoría padre no válida.', 'error')
            return redirect(url_for('admin.create_subcategory'))

        subcategoria = Subcategoria(
            subcategoria=subcategoria_nombre,
            id_categoria=id_categoria,
            ruta=ruta,
            estado=estado
        )

        db.session.add(subcategoria)
        db.session.commit()

        flash('Subcategoría creada exitosamente.', 'success')
        return redirect(url_for('admin.subcategories'))

    categorias = Categoria.query.filter_by(estado=1).all()
    return render_template('admin/subcategory_create.html', categorias=categorias)


@admin_bp.route('/subcategories/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_subcategory(id):
    """Edit subcategory."""
    subcategoria = Subcategoria.query.get_or_404(id)

    if request.method == 'POST':
        subcategoria_nombre = request.form.get('subcategoria')
        ruta = request.form.get('ruta')
        id_categoria = request.form.get('id_categoria', type=int)
        estado = request.form.get('estado', 1, type=int)

        if not subcategoria_nombre or not ruta or not id_categoria:
            flash('Todos los campos son requeridos.', 'error')
            return redirect(url_for('admin.edit_subcategory', id=id))

        # Check if ruta already exists (except current)
        existing = Subcategoria.query.filter(
            Subcategoria.ruta == ruta,
            Subcategoria.id != id
        ).first()
        if existing:
            flash('La ruta ya existe. Use una ruta única.', 'error')
            return redirect(url_for('admin.edit_subcategory', id=id))

        # Verify parent category exists
        categoria = Categoria.query.get(id_categoria)
        if not categoria:
            flash('Categoría padre no válida.', 'error')
            return redirect(url_for('admin.edit_subcategory', id=id))

        subcategoria.subcategoria = subcategoria_nombre
        subcategoria.ruta = ruta
        subcategoria.id_categoria = id_categoria
        subcategoria.estado = estado

        db.session.commit()

        flash('Subcategoría actualizada exitosamente.', 'success')
        return redirect(url_for('admin.subcategories'))

    categorias = Categoria.query.all()
    return render_template('admin/subcategory_edit.html',
                          subcategoria=subcategoria,
                          categorias=categorias)


@admin_bp.route('/subcategories/delete/<int:id>', methods=['POST'])
@admin_required
def delete_subcategory(id):
    """Delete subcategory."""
    subcategoria = Subcategoria.query.get_or_404(id)

    # Check if has products
    if subcategoria.get_products_count() > 0:
        flash('No se puede eliminar una subcategoría con productos asociados.', 'error')
        return redirect(url_for('admin.subcategories'))

    db.session.delete(subcategoria)
    db.session.commit()

    flash('Subcategoría eliminada exitosamente.', 'success')
    return redirect(url_for('admin.subcategories'))


@admin_bp.route('/subcategories/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_subcategory(id):
    """Toggle subcategory status."""
    subcategoria = Subcategoria.query.get_or_404(id)
    subcategoria.estado = 0 if subcategoria.estado == 1 else 1
    db.session.commit()

    return jsonify({'success': True, 'estado': subcategoria.estado})


# ===========================
# SLIDES MANAGEMENT
# ===========================

@admin_bp.route('/slides')
@admin_required
def slides():
    """Slides management page."""
    slides = Slide.query.order_by(Slide.orden.asc(), Slide.fecha.desc()).all()
    return render_template('admin/slides.html', slides=slides)


@admin_bp.route('/slides/create', methods=['GET', 'POST'])
@admin_required
def create_slide():
    """Create new slide."""
    if request.method == 'POST':
        from PIL import Image
        
        nombre = request.form.get('nombre')
        titulo1 = request.form.get('titulo1', '')
        titulo2 = request.form.get('titulo2', '')
        titulo3 = request.form.get('titulo3', '')
        boton = request.form.get('boton', '')
        url = request.form.get('url', '')
        orden = request.form.get('orden', 0, type=int)
        
        if not nombre:
            flash('El nombre es requerido.', 'error')
            return redirect(url_for('admin.create_slide'))
        
        # Handle background image upload
        img_fondo = ''
        if 'imgFondo' in request.files:
            file = request.files['imgFondo']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_folder = os.path.join('app/static/uploads/slides')
                
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)
                
                filepath = os.path.join(upload_folder, filename)
                
                # Save and optionally resize
                try:
                    img = Image.open(file)
                    # Resize to 1920x600 for slides
                    img = img.resize((1920, 600), Image.Resampling.LANCZOS)
                    img.save(filepath)
                except:
                    file.seek(0)
                    file.save(filepath)
                
                img_fondo = f'uploads/slides/{filename}'
        
        if not img_fondo:
            flash('La imagen de fondo es requerida.', 'error')
            return redirect(url_for('admin.create_slide'))
        
        slide = Slide(
            nombre=nombre,
            imgFondo=img_fondo,
            titulo1=titulo1,
            titulo2=titulo2,
            titulo3=titulo3,
            boton=boton,
            url=url,
            orden=orden
        )
        
        db.session.add(slide)
        db.session.commit()
        
        flash('Slide creado exitosamente.', 'success')
        return redirect(url_for('admin.slides'))
    
    return render_template('admin/slide_create.html')


@admin_bp.route('/slides/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_slide(id):
    """Edit slide."""
    slide = Slide.query.get_or_404(id)
    
    if request.method == 'POST':
        from PIL import Image
        
        slide.nombre = request.form.get('nombre')
        slide.titulo1 = request.form.get('titulo1', '')
        slide.titulo2 = request.form.get('titulo2', '')
        slide.titulo3 = request.form.get('titulo3', '')
        slide.boton = request.form.get('boton', '')
        slide.url = request.form.get('url', '')
        slide.orden = request.form.get('orden', 0, type=int)
        
        # Handle background image upload (optional on edit)
        if 'imgFondo' in request.files:
            file = request.files['imgFondo']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_folder = os.path.join('app/static/uploads/slides')
                
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)
                
                filepath = os.path.join(upload_folder, filename)
                
                try:
                    img = Image.open(file)
                    img = img.resize((1920, 600), Image.Resampling.LANCZOS)
                    img.save(filepath)
                except:
                    file.seek(0)
                    file.save(filepath)
                
                slide.imgFondo = f'uploads/slides/{filename}'
        
        db.session.commit()
        
        flash('Slide actualizado exitosamente.', 'success')
        return redirect(url_for('admin.slides'))
    
    return render_template('admin/slide_edit.html', slide=slide)


@admin_bp.route('/slides/delete/<int:id>', methods=['POST'])
@admin_required
def delete_slide(id):
    """Delete slide."""
    slide = Slide.query.get_or_404(id)

    db.session.delete(slide)
    db.session.commit()

    flash('Slide eliminado exitosamente.', 'success')
    return redirect(url_for('admin.slides'))


@admin_bp.route('/slides/reorder', methods=['POST'])
@admin_required
def reorder_slides():
    """Reorder slides via drag and drop."""
    try:
        data = request.get_json()
        slide_ids = data.get('slide_ids', [])

        if not slide_ids:
            return jsonify({'success': False, 'message': 'No slide IDs provided'}), 400

        # Update orden for each slide
        for index, slide_id in enumerate(slide_ids):
            slide = Slide.query.get(slide_id)
            if slide:
                slide.orden = index + 1

        db.session.commit()

        return jsonify({'success': True, 'message': 'Orden actualizado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# ===========================
# BANNERS MANAGEMENT
# ===========================

@admin_bp.route('/banners')
@admin_required
def banners():
    """Banners management page with DataTables."""
    categorias = Categoria.query.filter_by(estado=1).all()
    subcategorias = Subcategoria.query.filter_by(estado=1).all()
    return render_template('admin/banners.html',
                         categorias=categorias,
                         subcategorias=subcategorias)


@admin_bp.route('/banners/ajax')
@admin_required
def banners_ajax():
    """AJAX endpoint for DataTables banner listing."""
    # DataTables parameters
    draw = request.args.get('draw', type=int, default=1)
    start = request.args.get('start', type=int, default=0)
    length = request.args.get('length', type=int, default=10)
    search_value = request.args.get('search[value]', default='')

    # Base query
    query = Banner.query

    # Search filter
    if search_value:
        query = query.filter(
            db.or_(
                Banner.ruta.like(f'%{search_value}%'),
                Banner.tipo.like(f'%{search_value}%')
            )
        )

    # Total records
    total_records = Banner.query.count()
    filtered_records = query.count()

    # Get paginated data
    banners = query.order_by(Banner.fecha.desc()).offset(start).limit(length).all()

    # Format data for DataTables
    data = []
    for banner in banners:
        data.append({
            'id': banner.id,
            'img': banner.get_image_url(),
            'img_filename': banner.img,
            'tipo': banner.tipo,
            'ruta': banner.ruta,
            'estado': banner.estado,
            'activo': banner.is_active(),
            'fecha': banner.fecha.strftime('%d/%m/%Y') if banner.fecha else ''
        })

    return jsonify({
        'draw': draw,
        'recordsTotal': total_records,
        'recordsFiltered': filtered_records,
        'data': data
    })


@admin_bp.route('/banners/create', methods=['GET', 'POST'])
@admin_required
def create_banner():
    """Create new banner."""
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        ruta = request.form.get('ruta', 'general')
        estado = int(request.form.get('estado', 1))

        if not tipo:
            flash('El tipo de banner es requerido.', 'error')
            return redirect(url_for('admin.create_banner'))

        # Validate tipo
        if tipo not in ['categorias', 'subcategorias', 'general']:
            flash('Tipo de banner inválido.', 'error')
            return redirect(url_for('admin.create_banner'))

        # If general type, force ruta to 'general'
        if tipo == 'general':
            ruta = 'general'
        elif not ruta or ruta == 'general':
            flash('Debe seleccionar una categoría o subcategoría.', 'error')
            return redirect(url_for('admin.create_banner'))

        # Handle image upload
        img_filename = ''
        if 'img' in request.files:
            file = request.files['img']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_folder = os.path.join('app/static/uploads/banners')

                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)

                filepath = os.path.join(upload_folder, filename)

                # Resize banner image to standard size (1200x400)
                try:
                    img = Image.open(file)
                    img = img.resize((1200, 400), Image.Resampling.LANCZOS)
                    img.save(filepath)
                except Exception as e:
                    flash(f'Error al procesar imagen: {e}', 'error')
                    return redirect(url_for('admin.create_banner'))

                img_filename = filename

        if not img_filename:
            flash('La imagen del banner es requerida.', 'error')
            return redirect(url_for('admin.create_banner'))

        # Create banner
        banner = Banner(
            ruta=ruta,
            tipo=tipo,
            img=img_filename,
            estado=estado
        )

        db.session.add(banner)
        db.session.commit()

        flash('Banner creado exitosamente.', 'success')
        return redirect(url_for('admin.banners'))

    # GET request - show form
    categorias = Categoria.query.filter_by(estado=1).all()
    subcategorias = Subcategoria.query.filter_by(estado=1).all()
    return render_template('admin/banner_form.html',
                         banner=None,
                         categorias=categorias,
                         subcategorias=subcategorias)


@admin_bp.route('/banners/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_banner(id):
    """Edit banner."""
    banner = Banner.query.get_or_404(id)

    if request.method == 'POST':
        tipo = request.form.get('tipo')
        ruta = request.form.get('ruta', 'general')
        estado = int(request.form.get('estado', 1))

        if not tipo:
            flash('El tipo de banner es requerido.', 'error')
            return redirect(url_for('admin.edit_banner', id=id))

        # Validate tipo
        if tipo not in ['categorias', 'subcategorias', 'general']:
            flash('Tipo de banner inválido.', 'error')
            return redirect(url_for('admin.edit_banner', id=id))

        # If general type, force ruta to 'general'
        if tipo == 'general':
            ruta = 'general'
        elif not ruta or ruta == 'general':
            flash('Debe seleccionar una categoría o subcategoría.', 'error')
            return redirect(url_for('admin.edit_banner', id=id))

        banner.tipo = tipo
        banner.ruta = ruta
        banner.estado = estado

        # Handle image upload (optional on edit)
        if 'img' in request.files:
            file = request.files['img']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_folder = os.path.join('app/static/uploads/banners')

                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)

                filepath = os.path.join(upload_folder, filename)

                # Resize banner image to standard size (1200x400)
                try:
                    img = Image.open(file)
                    img = img.resize((1200, 400), Image.Resampling.LANCZOS)
                    img.save(filepath)
                    banner.img = filename
                except Exception as e:
                    flash(f'Error al procesar imagen: {e}', 'error')
                    return redirect(url_for('admin.edit_banner', id=id))

        db.session.commit()

        flash('Banner actualizado exitosamente.', 'success')
        return redirect(url_for('admin.banners'))

    # GET request - show form
    categorias = Categoria.query.filter_by(estado=1).all()
    subcategorias = Subcategoria.query.filter_by(estado=1).all()
    return render_template('admin/banner_form.html',
                         banner=banner,
                         categorias=categorias,
                         subcategorias=subcategorias)


@admin_bp.route('/banners/delete/<int:id>', methods=['POST'])
@admin_required
def delete_banner(id):
    """Delete banner."""
    banner = Banner.query.get_or_404(id)

    # Delete image file
    try:
        img_path = os.path.join('app/static/uploads/banners', banner.img)
        if os.path.exists(img_path):
            os.remove(img_path)
    except Exception as e:
        print(f'Error deleting banner image: {e}')

    db.session.delete(banner)
    db.session.commit()

    flash('Banner eliminado exitosamente.', 'success')
    return redirect(url_for('admin.banners'))


@admin_bp.route('/banners/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_banner(id):
    """Toggle banner status (activate/deactivate)."""
    try:
        banner = Banner.query.get_or_404(id)

        if banner.is_active():
            banner.deactivate()
        else:
            banner.activate()

        return jsonify({
            'success': True,
            'estado': banner.estado,
            'activo': banner.is_active()
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ===========================
# ADMINISTRATORS MANAGEMENT
# ===========================

@admin_bp.route('/administradores')
@admin_required
def administradores():
    """Administrators management page."""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')

    query = Administrador.query

    if search:
        query = query.filter(
            db.or_(
                Administrador.nombre.like(f'%{search}%'),
                Administrador.email.like(f'%{search}%')
            )
        )

    administradores = query.order_by(Administrador.fecha.desc()).paginate(
        page=page, per_page=25, error_out=False
    )

    return render_template('admin/administradores.html', administradores=administradores)


@admin_bp.route('/administradores/crear', methods=['GET', 'POST'])
@admin_required
def crear_administrador():
    """Create new administrator."""
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        password = request.form.get('password')
        perfil = request.form.get('perfil', 'editor')
        estado = int(request.form.get('estado', 1))

        if not nombre or not email or not password:
            flash('Nombre, email y contraseña son requeridos.', 'error')
            return redirect(url_for('admin.crear_administrador'))

        # Check if email already exists
        existing = Administrador.query.filter_by(email=email).first()
        if existing:
            flash('El email ya está registrado.', 'error')
            return redirect(url_for('admin.crear_administrador'))

        # Validate perfil
        if perfil not in ['administrador', 'editor']:
            flash('Perfil inválido.', 'error')
            return redirect(url_for('admin.crear_administrador'))

        # Handle photo upload
        foto = ''
        if 'foto' in request.files:
            file = request.files['foto']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_folder = os.path.join('app/static/uploads/admins')

                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)

                filepath = os.path.join(upload_folder, filename)

                # Resize to 300x300
                try:
                    img = Image.open(file)
                    img = img.resize((300, 300), Image.Resampling.LANCZOS)
                    img.save(filepath)
                    foto = f'uploads/admins/{filename}'
                except Exception as e:
                    flash(f'Error al procesar foto: {e}', 'error')
                    return redirect(url_for('admin.crear_administrador'))

        # Create administrator
        admin = Administrador(
            nombre=nombre,
            email=email,
            perfil=perfil,
            estado=estado,
            foto=foto
        )
        admin.set_password(password)

        db.session.add(admin)
        db.session.commit()

        flash(f'Administrador "{nombre}" creado exitosamente.', 'success')
        return redirect(url_for('admin.administradores'))

    return render_template('admin/administrador_form.html', administrador=None)


@admin_bp.route('/administradores/editar/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar_administrador(id):
    """Edit administrator."""
    administrador = Administrador.query.get_or_404(id)

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        password = request.form.get('password')
        perfil = request.form.get('perfil', 'editor')
        estado = int(request.form.get('estado', 1))

        if not nombre or not email:
            flash('Nombre y email son requeridos.', 'error')
            return redirect(url_for('admin.editar_administrador', id=id))

        # Check if email already exists (except current)
        existing = Administrador.query.filter(
            Administrador.email == email,
            Administrador.id != id
        ).first()
        if existing:
            flash('El email ya está registrado.', 'error')
            return redirect(url_for('admin.editar_administrador', id=id))

        # Validate perfil
        if perfil not in ['administrador', 'editor']:
            flash('Perfil inválido.', 'error')
            return redirect(url_for('admin.editar_administrador', id=id))

        # Update basic info
        administrador.nombre = nombre
        administrador.email = email
        administrador.perfil = perfil
        administrador.estado = estado

        # Update password if provided
        if password:
            administrador.set_password(password)

        # Handle photo upload (optional on edit)
        if 'foto' in request.files:
            file = request.files['foto']
            if file and file.filename:
                filename = secure_filename(file.filename)
                upload_folder = os.path.join('app/static/uploads/admins')

                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)

                filepath = os.path.join(upload_folder, filename)

                # Resize to 300x300
                try:
                    img = Image.open(file)
                    img = img.resize((300, 300), Image.Resampling.LANCZOS)
                    img.save(filepath)
                    administrador.foto = f'uploads/admins/{filename}'
                except Exception as e:
                    flash(f'Error al procesar foto: {e}', 'error')
                    return redirect(url_for('admin.editar_administrador', id=id))

        db.session.commit()

        flash(f'Administrador "{nombre}" actualizado exitosamente.', 'success')
        return redirect(url_for('admin.administradores'))

    return render_template('admin/administrador_form.html', administrador=administrador)


@admin_bp.route('/administradores/eliminar/<int:id>', methods=['POST'])
@admin_required
def eliminar_administrador(id):
    """Delete administrator."""
    administrador = Administrador.query.get_or_404(id)

    # Prevent deleting yourself
    if administrador.id == session.get('admin_id'):
        flash('No puedes eliminar tu propia cuenta.', 'error')
        return redirect(url_for('admin.administradores'))

    # Prevent deleting last administrator
    total_admins = Administrador.query.filter_by(perfil='administrador', estado=1).count()
    if administrador.perfil == 'administrador' and administrador.estado == 1 and total_admins <= 1:
        flash('No se puede eliminar el último administrador activo del sistema.', 'error')
        return redirect(url_for('admin.administradores'))

    nombre = administrador.nombre

    # Delete photo file
    if administrador.foto:
        try:
            foto_path = os.path.join('app/static', administrador.foto)
            if os.path.exists(foto_path):
                os.remove(foto_path)
        except Exception as e:
            print(f'Error deleting admin photo: {e}')

    db.session.delete(administrador)
    db.session.commit()

    flash(f'Administrador "{nombre}" eliminado exitosamente.', 'success')
    return redirect(url_for('admin.administradores'))


@admin_bp.route('/administradores/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_administrador(id):
    """Toggle administrator status."""
    try:
        administrador = Administrador.query.get_or_404(id)

        # Prevent deactivating yourself
        if administrador.id == session.get('admin_id'):
            return jsonify({'success': False, 'message': 'No puedes desactivar tu propia cuenta'}), 400

        # Prevent deactivating last active administrator
        if administrador.perfil == 'administrador' and administrador.estado == 1:
            total_admins = Administrador.query.filter_by(perfil='administrador', estado=1).count()
            if total_admins <= 1:
                return jsonify({
                    'success': False,
                    'message': 'No se puede desactivar el último administrador activo'
                }), 400

        # Toggle status
        administrador.estado = 0 if administrador.estado == 1 else 1
        db.session.commit()

        return jsonify({
            'success': True,
            'estado': administrador.estado
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ===========================
# COUPONS MANAGEMENT
# ===========================

@admin_bp.route('/coupons')
@admin_required
def coupons():
    """Coupons management page."""
    from app.models.coupon import Cupon

    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')

    query = Cupon.query

    if search:
        query = query.filter(Cupon.codigo.contains(search.upper()))

    cupones = query.order_by(Cupon.fecha.desc()).paginate(
        page=page, per_page=25, error_out=False
    )

    return render_template('admin/coupons.html', cupones=cupones)


@admin_bp.route('/coupons/create', methods=['GET', 'POST'])
@admin_required
def create_coupon():
    """Create new coupon."""
    from app.models.coupon import Cupon

    if request.method == 'POST':
        codigo = request.form.get('codigo', '').upper()
        tipo = request.form.get('tipo')
        valor = float(request.form.get('valor', 0))
        descripcion = request.form.get('descripcion', '')
        usos_maximos = int(request.form.get('usos_maximos', 0))
        monto_minimo = float(request.form.get('monto_minimo', 0))
        estado = int(request.form.get('estado', 1))

        if not codigo or not tipo or valor <= 0:
            flash('Código, tipo y valor son requeridos.', 'error')
            return redirect(url_for('admin.create_coupon'))

        # Validate tipo
        if tipo not in ['porcentaje', 'fijo']:
            flash('Tipo de cupón inválido.', 'error')
            return redirect(url_for('admin.create_coupon'))

        # Validate percentage value
        if tipo == 'porcentaje' and (valor < 0 or valor > 100):
            flash('El porcentaje debe estar entre 0 y 100.', 'error')
            return redirect(url_for('admin.create_coupon'))

        # Check if code already exists
        existing = Cupon.query.filter_by(codigo=codigo).first()
        if existing:
            flash('El código de cupón ya existe.', 'error')
            return redirect(url_for('admin.create_coupon'))

        # Handle dates
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')

        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%dT%H:%M')
        else:
            fecha_inicio = datetime.utcnow()

        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%dT%H:%M')
        else:
            fecha_fin = None

        cupon = Cupon(
            codigo=codigo,
            tipo=tipo,
            valor=valor,
            descripcion=descripcion,
            usos_maximos=usos_maximos,
            monto_minimo=monto_minimo,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estado=estado
        )

        db.session.add(cupon)
        db.session.commit()

        flash('Cupón creado exitosamente.', 'success')
        return redirect(url_for('admin.coupons'))

    return render_template('admin/coupon_form.html', cupon=None)


@admin_bp.route('/coupons/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_coupon(id):
    """Edit coupon."""
    from app.models.coupon import Cupon

    cupon = Cupon.query.get_or_404(id)

    if request.method == 'POST':
        codigo = request.form.get('codigo', '').upper()
        tipo = request.form.get('tipo')
        valor = float(request.form.get('valor', 0))
        descripcion = request.form.get('descripcion', '')
        usos_maximos = int(request.form.get('usos_maximos', 0))
        monto_minimo = float(request.form.get('monto_minimo', 0))
        estado = int(request.form.get('estado', 1))

        if not codigo or not tipo or valor <= 0:
            flash('Código, tipo y valor son requeridos.', 'error')
            return redirect(url_for('admin.edit_coupon', id=id))

        # Validate tipo
        if tipo not in ['porcentaje', 'fijo']:
            flash('Tipo de cupón inválido.', 'error')
            return redirect(url_for('admin.edit_coupon', id=id))

        # Validate percentage value
        if tipo == 'porcentaje' and (valor < 0 or valor > 100):
            flash('El porcentaje debe estar entre 0 y 100.', 'error')
            return redirect(url_for('admin.edit_coupon', id=id))

        # Check if code already exists (except current)
        existing = Cupon.query.filter(Cupon.codigo == codigo, Cupon.id != id).first()
        if existing:
            flash('El código de cupón ya existe.', 'error')
            return redirect(url_for('admin.edit_coupon', id=id))

        # Handle dates
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')

        if fecha_inicio:
            cupon.fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%dT%H:%M')

        if fecha_fin:
            cupon.fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%dT%H:%M')
        else:
            cupon.fecha_fin = None

        cupon.codigo = codigo
        cupon.tipo = tipo
        cupon.valor = valor
        cupon.descripcion = descripcion
        cupon.usos_maximos = usos_maximos
        cupon.monto_minimo = monto_minimo
        cupon.estado = estado

        db.session.commit()

        flash('Cupón actualizado exitosamente.', 'success')
        return redirect(url_for('admin.coupons'))

    return render_template('admin/coupon_form.html', cupon=cupon)


@admin_bp.route('/coupons/delete/<int:id>', methods=['POST'])
@admin_required
def delete_coupon(id):
    """Delete coupon."""
    from app.models.coupon import Cupon

    cupon = Cupon.query.get_or_404(id)

    db.session.delete(cupon)
    db.session.commit()

    flash('Cupón eliminado exitosamente.', 'success')
    return redirect(url_for('admin.coupons'))


@admin_bp.route('/coupons/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_coupon(id):
    """Toggle coupon status."""
    from app.models.coupon import Cupon

    cupon = Cupon.query.get_or_404(id)
    cupon.estado = 0 if cupon.estado == 1 else 1
    db.session.commit()

    return jsonify({'success': True, 'estado': cupon.estado})


# ===========================
# Comments Management
# ===========================

@admin_bp.route('/comments')
@admin_required
def comments():
    """Comments management page."""
    # Get filter parameters
    estado_filter = request.args.get('estado', '')
    calificacion_filter = request.args.get('calificacion', '')
    producto_filter = request.args.get('producto', '')
    search_query = request.args.get('search', '')

    # Base query with relationships
    query = Comentario.query

    # Apply filters
    if estado_filter:
        query = query.filter_by(estado=estado_filter)

    if calificacion_filter:
        query = query.filter(Comentario.calificacion >= float(calificacion_filter))

    if producto_filter:
        query = query.filter_by(id_producto=int(producto_filter))

    if search_query:
        query = query.join(Comentario.usuario).filter(
            db.or_(
                Comentario.comentario.ilike(f'%{search_query}%'),
                User.nombre.ilike(f'%{search_query}%')
            )
        )

    # Order by date descending
    comentarios = query.order_by(Comentario.fecha.desc()).all()

    # Get statistics
    total_comentarios = Comentario.query.count()
    pendientes = Comentario.query.filter_by(estado=Comentario.ESTADO_PENDIENTE).count()
    aprobados = Comentario.query.filter_by(estado=Comentario.ESTADO_APROBADO).count()
    rechazados = Comentario.query.filter_by(estado=Comentario.ESTADO_RECHAZADO).count()

    # Get products for filter dropdown
    productos = Producto.query.filter_by(estado=1).order_by(Producto.titulo).all()

    return render_template('admin/comments.html',
                         comentarios=comentarios,
                         total_comentarios=total_comentarios,
                         pendientes=pendientes,
                         aprobados=aprobados,
                         rechazados=rechazados,
                         productos=productos,
                         estado_filter=estado_filter,
                         calificacion_filter=calificacion_filter,
                         producto_filter=producto_filter,
                         search_query=search_query)


@admin_bp.route('/comments/approve/<int:id>', methods=['POST'])
@admin_required
def approve_comment(id):
    """Approve comment."""
    comentario = Comentario.query.get_or_404(id)
    comentario.aprobar()

    flash(f'Comentario de {comentario.usuario.nombre} aprobado.', 'success')
    return redirect(url_for('admin.comments'))


@admin_bp.route('/comments/reject/<int:id>', methods=['POST'])
@admin_required
def reject_comment(id):
    """Reject comment."""
    comentario = Comentario.query.get_or_404(id)
    comentario.rechazar()

    flash(f'Comentario de {comentario.usuario.nombre} rechazado.', 'success')
    return redirect(url_for('admin.comments'))


@admin_bp.route('/comments/delete/<int:id>', methods=['POST'])
@admin_required
def delete_comment(id):
    """Delete comment."""
    comentario = Comentario.query.get_or_404(id)
    usuario_nombre = comentario.usuario.nombre

    db.session.delete(comentario)
    db.session.commit()

    flash(f'Comentario de {usuario_nombre} eliminado permanentemente.', 'success')
    return redirect(url_for('admin.comments'))


@admin_bp.route('/comments/respond/<int:id>', methods=['POST'])
@admin_required
def respond_comment(id):
    """Respond to comment as admin."""
    comentario = Comentario.query.get_or_404(id)
    respuesta = request.form.get('respuesta', '').strip()

    if not respuesta:
        flash('La respuesta no puede estar vacía.', 'error')
        return redirect(url_for('admin.comments'))

    comentario.respuesta_admin = respuesta
    comentario.fecha_moderacion = datetime.utcnow()

    # Auto-approve when responding
    if comentario.estado == Comentario.ESTADO_PENDIENTE:
        comentario.estado = Comentario.ESTADO_APROBADO

    db.session.commit()

    flash(f'Respuesta publicada en el comentario de {comentario.usuario.nombre}.', 'success')
    return redirect(url_for('admin.comments'))


@admin_bp.route('/comments/toggle/<int:id>', methods=['POST'])
@admin_required
def toggle_comment(id):
    """Toggle comment status (approve/reject)."""
    comentario = Comentario.query.get_or_404(id)

    if comentario.estado == Comentario.ESTADO_APROBADO:
        comentario.rechazar()
        nuevo_estado = 'rechazado'
    else:
        comentario.aprobar()
        nuevo_estado = 'aprobado'

    return jsonify({
        'success': True,
        'estado': comentario.estado,
        'estado_display': nuevo_estado
    })


# ===========================
# STORE CUSTOMIZATION
# ===========================

@admin_bp.route('/personalizacion', methods=['GET', 'POST'])
@admin_required
def personalizacion():
    """Store customization page (logo, colors, favicon, social media, analytics)."""
    plantilla = Plantilla.get_settings()

    if request.method == 'POST':
        try:
            # Basic colors and text
            plantilla.colorFondo = request.form.get('colorFondo', '#ffffff')
            plantilla.colorTexto = request.form.get('colorTexto', '#000000')
            plantilla.barraSuperior = request.form.get('barraSuperior', '')
            plantilla.textoSuperior = request.form.get('textoSuperior', '')

            # Analytics and tracking
            plantilla.pixelFacebook = request.form.get('pixelFacebook', '')
            plantilla.googleAnalytics = request.form.get('googleAnalytics', '')
            plantilla.apiFacebook = request.form.get('apiFacebook', '')

            # Social Networks
            redes = {
                'facebook': request.form.get('facebook', ''),
                'instagram': request.form.get('instagram', ''),
                'twitter': request.form.get('twitter', ''),
                'youtube': request.form.get('youtube', ''),
                'linkedin': request.form.get('linkedin', ''),
                'tiktok': request.form.get('tiktok', ''),
                'whatsapp': request.form.get('whatsapp', '')
            }
            plantilla.set_social_networks(redes)

            # Handle logo upload
            if 'logo' in request.files:
                file = request.files['logo']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    upload_folder = os.path.join('app/static/uploads/branding')

                    if not os.path.exists(upload_folder):
                        os.makedirs(upload_folder)

                    filepath = os.path.join(upload_folder, f'logo_{filename}')

                    # Resize logo to max width 400px maintaining aspect ratio
                    try:
                        img = Image.open(file)
                        max_width = 400
                        if img.width > max_width:
                            ratio = max_width / img.width
                            new_height = int(img.height * ratio)
                            img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
                        img.save(filepath)
                        plantilla.logo = f'uploads/branding/logo_{filename}'
                    except Exception as e:
                        flash(f'Error al procesar logo: {e}', 'error')

            # Handle favicon upload
            if 'icono' in request.files:
                file = request.files['icono']
                if file and file.filename:
                    filename = secure_filename(file.filename)
                    upload_folder = os.path.join('app/static/uploads/branding')

                    if not os.path.exists(upload_folder):
                        os.makedirs(upload_folder)

                    filepath = os.path.join(upload_folder, f'favicon_{filename}')

                    # Resize favicon to 32x32
                    try:
                        img = Image.open(file)
                        img = img.resize((32, 32), Image.Resampling.LANCZOS)
                        img.save(filepath)
                        plantilla.icono = f'uploads/branding/favicon_{filename}'
                    except Exception as e:
                        flash(f'Error al procesar favicon: {e}', 'error')

            db.session.commit()
            flash('Personalización actualizada exitosamente!', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al guardar personalización: {e}', 'error')

        return redirect(url_for('admin.personalizacion'))

    # GET request - show form
    redes = plantilla.get_social_networks()

    return render_template('admin/personalizacion.html',
                         plantilla=plantilla,
                         redes=redes)


# ==================== MESSAGING SYSTEM ====================

@admin_bp.route('/mensajes')
@admin_required
def mensajes():
    """Admin inbox - received messages."""
    admin = Administrador.query.get(session['admin_id'])

    # Get all received messages
    mensajes = Mensaje.query.filter_by(
        destinatario_tipo='admin',
        destinatario_id=admin.id
    ).order_by(Mensaje.fecha.desc()).all()

    # Count unread messages
    no_leidos = Mensaje.contar_no_leidos('admin', admin.id)

    return render_template('admin/mensajes.html',
                         mensajes=mensajes,
                         no_leidos=no_leidos,
                         tab='recibidos')


@admin_bp.route('/mensajes/enviados')
@admin_required
def mensajes_enviados():
    """Admin sent messages."""
    admin = Administrador.query.get(session['admin_id'])

    # Get all sent messages
    mensajes = Mensaje.query.filter_by(
        remitente_tipo='admin',
        remitente_id=admin.id
    ).order_by(Mensaje.fecha.desc()).all()

    return render_template('admin/mensajes.html',
                         mensajes=mensajes,
                         tab='enviados')


@admin_bp.route('/mensajes/nuevo', methods=['GET', 'POST'])
@admin_required
def nuevo_mensaje():
    """Compose new message."""
    admin = Administrador.query.get(session['admin_id'])

    if request.method == 'POST':
        try:
            destinatario_id = request.form.get('destinatario_id')
            asunto = request.form.get('asunto', '').strip()
            contenido = request.form.get('contenido', '').strip()

            if not all([destinatario_id, asunto, contenido]):
                flash('Todos los campos son obligatorios.', 'error')
                return redirect(url_for('admin.nuevo_mensaje'))

            # Send message (admin to user)
            mensaje = Mensaje.enviar_mensaje(
                remitente_tipo='admin',
                remitente_id=admin.id,
                destinatario_tipo='user',
                destinatario_id=int(destinatario_id),
                asunto=asunto,
                contenido=contenido
            )

            flash(f'Mensaje enviado exitosamente a {mensaje.get_destinatario_nombre()}.', 'success')
            return redirect(url_for('admin.mensajes_enviados'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al enviar mensaje: {e}', 'error')

    # GET - show form
    usuarios = User.query.order_by(User.nombre).all()
    return render_template('admin/mensaje_form.html',
                         usuarios=usuarios,
                         mensaje=None)


@admin_bp.route('/mensajes/<int:id>')
@admin_required
def ver_mensaje(id):
    """View message details."""
    admin = Administrador.query.get(session['admin_id'])
    mensaje = Mensaje.query.get_or_404(id)

    # Verify admin has access to this message (either sender or recipient)
    if not ((mensaje.destinatario_tipo == 'admin' and mensaje.destinatario_id == admin.id) or
            (mensaje.remitente_tipo == 'admin' and mensaje.remitente_id == admin.id)):
        flash('No tiene permiso para ver este mensaje.', 'error')
        return redirect(url_for('admin.mensajes'))

    # Mark as read if admin is recipient
    if mensaje.destinatario_tipo == 'admin' and mensaje.destinatario_id == admin.id:
        mensaje.marcar_como_leido()

    # Get conversation thread
    if mensaje.mensaje_padre_id:
        # This is a reply, get parent and all siblings
        padre = Mensaje.query.get(mensaje.mensaje_padre_id)
        conversacion = [padre] + list(padre.respuestas.all())
    else:
        # This is a parent message, get all replies
        conversacion = [mensaje] + list(mensaje.respuestas.all())

    return render_template('admin/mensaje_detalle.html',
                         mensaje=mensaje,
                         conversacion=conversacion)


@admin_bp.route('/mensajes/<int:id>/responder', methods=['GET', 'POST'])
@admin_required
def responder_mensaje(id):
    """Reply to a message."""
    admin = Administrador.query.get(session['admin_id'])
    mensaje_original = Mensaje.query.get_or_404(id)

    # Verify admin has access
    if not ((mensaje_original.destinatario_tipo == 'admin' and mensaje_original.destinatario_id == admin.id) or
            (mensaje_original.remitente_tipo == 'admin' and mensaje_original.remitente_id == admin.id)):
        flash('No tiene permiso para responder este mensaje.', 'error')
        return redirect(url_for('admin.mensajes'))

    if request.method == 'POST':
        try:
            contenido = request.form.get('contenido', '').strip()

            if not contenido:
                flash('El contenido del mensaje es obligatorio.', 'error')
                return redirect(url_for('admin.responder_mensaje', id=id))

            # Determine recipient (whoever is NOT the current admin)
            if mensaje_original.remitente_tipo == 'admin' and mensaje_original.remitente_id == admin.id:
                # Admin sent original, reply to recipient
                dest_tipo = mensaje_original.destinatario_tipo
                dest_id = mensaje_original.destinatario_id
            else:
                # Admin received original, reply to sender
                dest_tipo = mensaje_original.remitente_tipo
                dest_id = mensaje_original.remitente_id

            # Find the root message for threading
            mensaje_padre_id = mensaje_original.mensaje_padre_id if mensaje_original.mensaje_padre_id else mensaje_original.id

            # Send reply
            respuesta = Mensaje.enviar_mensaje(
                remitente_tipo='admin',
                remitente_id=admin.id,
                destinatario_tipo=dest_tipo,
                destinatario_id=dest_id,
                asunto=f"Re: {mensaje_original.asunto}",
                contenido=contenido,
                mensaje_padre_id=mensaje_padre_id
            )

            flash('Respuesta enviada exitosamente.', 'success')
            return redirect(url_for('admin.ver_mensaje', id=mensaje_padre_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error al enviar respuesta: {e}', 'error')

    # GET - show reply form
    return render_template('admin/mensaje_form.html',
                         mensaje=mensaje_original,
                         es_respuesta=True)


@admin_bp.route('/mensajes/<int:id>/eliminar', methods=['POST'])
@admin_required
def eliminar_mensaje(id):
    """Delete a message."""
    admin = Administrador.query.get(session['admin_id'])
    mensaje = Mensaje.query.get_or_404(id)

    # Verify admin has access
    if not ((mensaje.destinatario_tipo == 'admin' and mensaje.destinatario_id == admin.id) or
            (mensaje.remitente_tipo == 'admin' and mensaje.remitente_id == admin.id)):
        flash('No tiene permiso para eliminar este mensaje.', 'error')
        return redirect(url_for('admin.mensajes'))

    try:
        asunto = mensaje.asunto
        db.session.delete(mensaje)
        db.session.commit()
        flash(f'Mensaje "{asunto}" eliminado exitosamente.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar mensaje: {e}', 'error')

    return redirect(url_for('admin.mensajes'))


@admin_bp.route('/mensajes/marcar-leido/<int:id>', methods=['POST'])
@admin_required
def marcar_mensaje_leido(id):
    """Mark message as read via AJAX."""
    admin = Administrador.query.get(session['admin_id'])
    mensaje = Mensaje.query.get_or_404(id)

    # Verify admin is recipient
    if mensaje.destinatario_tipo == 'admin' and mensaje.destinatario_id == admin.id:
        mensaje.marcar_como_leido()
        return jsonify({'success': True, 'message': 'Mensaje marcado como leído'})
    else:
        return jsonify({'success': False, 'message': 'No autorizado'}), 403
